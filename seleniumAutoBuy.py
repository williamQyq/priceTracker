from contextlib import nullcontext
from os import stat
from pandas._config import config
from pandas.core.base import NoNewAttributesMixin
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import datetime

import config
import pandas as pd
from selenium.webdriver.chrome.options import Options

from openpyxl import load_workbook

def main():
    
    df = init_excel()
    driver=init_chrome_driver() 

    iterate_quests(df,driver)

    print("[Status]: === Finished, output result===\n")

    driver.close()



def init_excel():
    prod_tracker = pd.read_excel('Tracker_Product.xlsx', sheet_name='TrackProduct')

    df = pd.DataFrame(prod_tracker)
    print(df)
    return df



def init_chrome_driver():
    user_agent = config.Config_static_user()
    user_agent_headers = user_agent.get_headers()

    chrome_options = Options() 
    chrome_options.add_experimental_option("detach",True)
    chrome_options.add_argument(f'user-agent={user_agent_headers}')
    driver = webdriver.Chrome(executable_path=user_agent.get_path(), options=chrome_options)

    
    print(f"[User Agent: \n{driver.execute_script('return navigator.userAgent;')}\n\n")

    return driver


def iterate_quests(df,driver):
    for index, row in df.iterrows():
        #status: nan; need quantity, else not
        status = pd.isnull(row.iloc[config.STATUS_INDEX])
        needQuantity = row.iloc[config.NEEDQUANTITY_INDEX]
        url = row.iloc[config.URL_INDEX]
        order_count = 0
        error = False

        if status:
            for i in range(needQuantity):  
                 
                try:
                    driver.get(url)
                    info = get_info(driver)
                    auto_purchase(driver)
                    order_count+=1
                    driver.delete_all_cookies()
                    print(f"[Success] *** Successful Placed Order No.{index} -- {i+1} times *** \n")
      
                except:
                    error = True
                    print(f"[Failure]: *** Failed auto purchase No.{index} -- {i+1} times ***\n")
                    record_result(config.filename,index,order_count,info,"fail")
                    driver.delete_all_cookies()
                    pass

            #write order_count to result excel
            if not error:    
                record_result(config.filename,index,order_count,info,"pass")
        else:
            print("[Status]: === Finished, empty result===\n")


def get_info(driver):
    info={
        "title":"",
        "price":""
    }
    try:
        element = driver.find_element_by_class_name("sku-title")
        title = element.find_element_by_tag_name('h1')

        element = driver.find_element_by_class_name("priceView-customer-price")
        price = element.find_element_by_tag_name('span')
        
        info.update({
            "title":title.text,
            "price":price.text
            })
    except:
        print("[Error]: === unable to get url title & price ===")

    return info

def record_result(filename,index,order_count,info,status):
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    current_time_formatted = str(datetime.datetime.now().strftime('%d-%b-%Y %Hh%Mm'))

    #record "Order" table: status column
    order_tab = ws.tables["Order"]
    ws[f"F{index+2}"]=status
    order_tab.ref = f"A1:F{index+2}"

    #record "Result" table: title, price, order_count, status, time
    result_tab = ws.tables["Result"]
    ws[f"I{index+2}"]=info["title"]
    ws[f"J{index+2}"]=info["price"]
    ws[f"K{index+2}"]=order_count
    ws[f"L{index+2}"]=status
    ws[f"H{index+2}"]=current_time_formatted

    result_tab.ref = f"H1:L{index+2}"

    wb.save(filename)


def auto_purchase(driver):

    user = config.Config_static_user()
    #click add to cart button
    try:
        add_to_cart_btn = driver.find_element_by_class_name('add-to-cart-button')
        add_to_cart_btn.click()
        print(f"[Status]: === purchase Order start, adding to cart === ")
    except:
        return print("[Failure]: === unable to add to cart ===")

    add_to_cart(driver,user)

    fill_personal_info(driver,user)

    fill_payment(driver,user)

    place_order(driver)

    return True



def add_to_cart(driver,user):

    #click a href www.bestbuy.com/cart
    try:
        element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/cart']")))
        # link = driver.find_element_by_link_text("Go to Cart")
        element.click()
    except:
        return print(f"[Error--Go to Cart]: ===Failure unable to wait for element loaded ===\n\n")
    
    try:
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "shop-alert")))
        return print("[Error] === Fail, duplicate product process === \n")
    except:
        pass


    #remove addons
    try:
        # element = WebDriverWait(driver,20).until(
        #     EC.element_to_be_clickable((By.LINK_TEXT, "Remove")))
        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "addon-base__actions")))
        addon_action_element = WebDriverWait(element,20).until(
            EC.element_to_be_clickable((By.LINK_TEXT, "Remove")))
        addon_action_element.click()
    except:
        print(f"[Error--Remove addons]: ===Failure unable to wait for remove element loaded ===:\n\n")
        pass

    #click checkout btn
    try:
        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "checkout-buttons__checkout")))
        element.click()
    except:
        return print(f"[Error--checkout-buttons__checkout]: ===Failure unable to wait for element loaded ===\n\n")
    
    # Login as existing customer
    try:
        login_account(driver,user)
    except:
        login_guest(driver)


    return print(f"[Status]: === Successful click loaded element ===\n")

def login_account(driver,user):
    try:
        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type='email']")))
        element.send_keys(user.get_account())
    except:
        print(f"[Error--input account]: ===Failure unable to wait for element loaded ===\n\n")
        return False

    try:
        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type='password']")))
        element.send_keys(user.get_account_password())
    except:
        print(f"[Error--input account password]: ===Failure unable to wait for element loaded ===\n\n")
        return False

    try:
        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "cia-form__controls")))
        element.click()
    except:
        print(f"[Error--input account]: ===Failure unable to wait for element loaded ===\n\n")
        return False

    try:
        element = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "cia-alert")))
        print(f"[Error--account or password incorrect]: ===Failure unable to login account ===\n\n")
        return False
    except:
        # click switch to shipping link
        try:
            element = WebDriverWait(driver,20).until(
                EC.element_to_be_clickable((By.LINK_TEXT, "Switch to Shipping")))
            element.click()
        except:
            print(f"[Error--switch free shipping]: ===Failure unable to wait for element loaded ===\n\n")
            return False

def login_guest(driver):
      # Or click continue as Guest
    try:
        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "cia-guest-content__continue")))
        element.click()
    except:
        return print(f"[Error--cia-guest-content__continue]: ===Failure unable to wait for element loaded ===\n\n")

    # click switch to shipping link
    try:
        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.LINK_TEXT, "Switch to Shipping")))
        element.click()
    except:
        return print(f"[Error--switch free shipping]: ===Failure unable to wait for element loaded ===\n\n")

def fill_personal_info(driver,user):
    
    try:
        element = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.ID,"sc-store-availability-modal")))
        print("[Status]: === Successful click sc-store modal ===\n")
        
        sel_store_element = WebDriverWait(element,20).until(
            EC.element_to_be_clickable((By.XPATH,"//div[@data-test-id = 'select-store-button']")))
        sel_store_element.click()
        print("[Status]: === Successful click select this location ===\n")

        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.LINK_TEXT, "ispu-card__switch")))
        element.click()
    
    except:
        print(f"[Attention --Select This Location Jumped]: === Jump select this location elements ===\n")

    # fill in contact info form
    try:
        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id$='firstName']")))
        element.send_keys(user.get_fname())
    except:
        return print(f"[Error--input firstName]: ===Failure unable to wait for element loaded ===\n\n")

    element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id$='lastName']")))
    element.send_keys(user.get_lname())

    element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id$='street']")))
    element.send_keys(user.get_address())

    element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id$='city']")))
    element.send_keys(user.get_city())

    element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.NAME, "state")))
    for option in element.find_elements_by_tag_name('option'):
        if option.text == f'{user.get_state()}':
            option.click()
            break

    element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id$='zipcode']")))
    element.send_keys(user.get_zipCode())

    try:
        element = WebDriverWait(driver,5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id$='emailAddress']")))
        element.send_keys(user.get_email())
    except:
        print(f"[Attention --input emailaddress Jumped]: === Jump input email, already login ===\n")

    element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id$='phone']")))
    element.send_keys(user.get_phone())

    try:
        # click Continue to payment information
        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "button--continue")))
        element.click()
        print(f"[Status]: === Successful fill personal info ===\n")
    except:
        print(f"[Error--continue to paymnent]: ===Failure unable to wait for element loaded ===\n\n")



def fill_payment(driver,user):

    try:
        element = WebDriverWait(driver,20).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME, "order-errors")))
        print(f"[Attention --Order shipping options changed]: === Jump to button-continue ===\n")
        continue_element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.CLASS_NAME, "button--continue")))
        continue_element.click()
    except:
        pass

    element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id$='optimized-cc-card-number']")))
    element.send_keys(user.get_credit_card_number())

    element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.NAME, "expiration-month")))
    element.send_keys(user.get_credit_card_expire_month())

    element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.NAME, "expiration-year")))
    element.send_keys(user.get_credit_card_expire_year())

    element = WebDriverWait(driver,20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id$='credit-card-cvv']")))
    element.send_keys(user.get_credit_card_cvv())

    print(f"[Status]: === Successful fill payment ===\n")

    return True



def place_order(driver):
    try:
        element = WebDriverWait(driver,20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "button--place-order")))
        #element.click()

        return True
    except:
        return False

main()